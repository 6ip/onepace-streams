import os
import re
import json
import time
import requests
import openpyxl
from bs4 import BeautifulSoup

SHEET_EXPORT_URL = "https://docs.google.com/spreadsheets/d/1HQRMJgu_zArp-sLnvFMDzOyjdsht87eFLECxMK858lA/export?format=xlsx"
LOCAL_EXCEL_FILE = "one_pace.xlsx"
TRACKER_FILE = "tracker.json"

PREFIX_MAP = {
    "Romance Dawn": "RO",
    "Orange Town": "OR",
    "Syrup Village": "SY",
    "Gaimon": "GA",
    "Baratie": "BA",
    "Arlong Park": "AR",
    "The Adventures of Buggys Crew": "BUGGYS_CREW",
    "Loguetown": "LO",
    "Reverse Mountain": "RM",
    "Whisky Peak": "WH",
    "The Trials of Koby-Meppo": "COVER_KOBYMEPPO",
    "Little Garden": "LI",
    "Drum Island": "DI",
    "Alabasta": "AL",
    "Jaya": "JA",
    "Skypiea": "SK",
    "Long Ring Long Land": "LR",
    "Water Seven": "WS",
    "Enies Lobby": "EN",
    "Post-Enies Lobby": "PEN",
    "Thriller Bark": "TB",
    "Sabaody Archipelago": "SAB",
    "Amazon Lily": "AM",
    "Impel Down": "IM",
    "The Adventures of the Straw Hat": "COVER_SHSS",
    "Marineford": "MA",
    "Post-War": "PW",
    "Return to Sabaody": "RTS",
    "Fishman Island": "FI",
    "Punk Hazard": "PH",
    "Dressrosa": "DR",
    "Zou": "ZO",
    "Whole Cake Island": "WC",
    "Reverie": "REV",
    "Wano": "WA",
    "Egghead": "EH"
}

def download_excel_file(url, filename):
    print(f"Downloading latest spreadsheet to {filename}...")
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()
        with open(filename, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        print("Download complete!\n")
        return True
    except Exception as e:
        print(f"Failed to download spreadsheet: {e}")
        return False

def get_info_hash(nyaa_url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        response = requests.get(nyaa_url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        magnet_tag = soup.find('a', href=re.compile(r'^magnet:\?xt=urn:btih:'))
        if magnet_tag:
            match = re.search(r'urn:btih:([a-zA-Z0-9]{40})', magnet_tag['href'])
            if match:
                return match.group(1).lower()
    except Exception as e:
        print(f"  [!] Error connecting to Nyaa: {e}")
    return None

def get_expected_filename(ep_name, arc_name):
    prefix = PREFIX_MAP.get(arc_name, arc_name[:2].upper())
    ep_name_str = str(ep_name).strip()
    
    match = re.search(r'(\d+)\s*$', ep_name_str)
    if not match:
        match = re.search(r'\b(\d{1,3})\b', ep_name_str)
        
    if match:
        ep_num = match.group(1)
    else:
        ep_num = "1"
        
    # Convert to integer to automatically drop leading zeros (e.g., "01" becomes 1)
    ep_num_int = int(ep_num)
    
    # If the arc is Long Ring Long Land (LR), shift the number down by 1
    if prefix == "LR":
        ep_num_int -= 1
        
    # Convert back to a string for the filename
    ep_num = str(ep_num_int)
        
    return f"{prefix}_{ep_num}.json"

def save_tracker(tracker_data):
    with open(TRACKER_FILE, 'w') as f:
        json.dump(tracker_data, f, indent=2)

def main():
    # Start the performance timer!
    start_time = time.time()
    new_files_count = 0 
    
    output_dir = "stream"
    os.makedirs(output_dir, exist_ok=True)
    
    tracker_data = {}
    if os.path.exists(TRACKER_FILE):
        with open(TRACKER_FILE, 'r') as f:
            tracker_data = json.load(f)
            
    success = download_excel_file(SHEET_EXPORT_URL, LOCAL_EXCEL_FILE)
    if not success:
        return

    print("Loading local spreadsheet...")
    workbook = openpyxl.load_workbook(LOCAL_EXCEL_FILE)
    
    for target_sheet in PREFIX_MAP.keys():
        if target_sheet not in workbook.sheetnames:
            continue
            
        sheet = workbook[target_sheet]
        print(f"\n--- Checking Arc: {target_sheet} ---")
        
        ep_col_idx, header_row = None, None
        
        for row in range(1, 10):
            for col in range(1, sheet.max_column + 1):
                cell_val = str(sheet.cell(row=row, column=col).value).strip()
                if "One Pace Episode" in cell_val: 
                    ep_col_idx = col
                    header_row = row
                    break
            if ep_col_idx:
                break
                
        if not ep_col_idx:
            print(f"  [!] Could not find the Episode Name column. Skipping.")
            continue

        for row in range(header_row + 1, sheet.max_row + 1):
            ep_name = sheet.cell(row=row, column=ep_col_idx).value
            
            if not ep_name:
                continue
                
            if "(G8)" in str(ep_name):
                print(f"  [~] Ignored {ep_name} (Skipping G8 filler)")
                continue
            
            filename = get_expected_filename(ep_name, target_sheet)
            filepath = os.path.join(output_dir, filename)
            
            nyaa_url = None
            for col in range(1, sheet.max_column + 1):
                if col == ep_col_idx: 
                    continue 
                    
                cell = sheet.cell(row=row, column=col)
                
                if cell.hyperlink and cell.hyperlink.target:
                    target = cell.hyperlink.target
                    if "nyaa.si" in target or "magnet:" in target:
                        nyaa_url = target
                        break
                        
                if cell.value and isinstance(cell.value, str) and 'HYPERLINK' in cell.value:
                    match = re.search(r'HYPERLINK\("([^"]+)"', cell.value)
                    if match:
                        target = match.group(1)
                        if "nyaa.si" in target or "magnet:" in target:
                            nyaa_url = target
                            break

            if not nyaa_url:
                print(f"  [-] No Nyaa link found for {ep_name}")
                continue

            if tracker_data.get(filename) == nyaa_url and os.path.exists(filepath):
                print(f"  [~] Skipped {filename} (Already up-to-date)")
                continue
                
            print(f"  [*] Processing {filename} (New or Updated!)")
            
            info_hash = get_info_hash(nyaa_url)
            if info_hash:
                data = {"streams": [{"infoHash": info_hash}]}
                with open(filepath, 'w') as f:
                    json.dump(data, f, indent=2)
                
                tracker_data[filename] = nyaa_url
                save_tracker(tracker_data)
                
                print(f"  [+] Saved {filename}")
                new_files_count += 1
                time.sleep(1) 
            else:
                print(f"  [-] Failed to get infoHash for {filename}")

    # Calculate final performance
    end_time = time.time()
    total_time = round(end_time - start_time, 2)
    
    print("\n=========================================")
    print(" ✅ SCRIPT FINISHED ")
    print("=========================================")
    print(f" ⏱️  Time taken: {total_time} seconds")
    print(f" 📥 New files downloaded: {new_files_count}")
    print("=========================================\n")

if __name__ == "__main__":
    main()