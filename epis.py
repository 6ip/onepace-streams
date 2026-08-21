import re
import time
import datetime
import requests
import openpyxl
import json
import csv
import urllib.request
import os
import sys
from io import StringIO

# --- Configuration ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_JSON = os.path.join(BASE_DIR, 'meta', 'pp_onepacee.json')

# --- Load Central Config ---
CONFIG_PATH = os.path.join(BASE_DIR, 'config.json')
with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
    CONFIG = json.load(f)
ARC_PREFIXES = CONFIG["ARC_MAP"]
TOTAL_SEASONS = CONFIG["TOTAL_SEASONS"]

# --- Spreadsheet source (must come AFTER CONFIG is loaded) ---
EPISODE_SHEET_ID = CONFIG.get("EPISODE_SHEET_ID", "1HQRMJgu_zArp-sLnvFMDzOyjdsht87eFLECxMK858lA")
EPISODE_XLSX_URL = f"https://docs.google.com/spreadsheets/d/{EPISODE_SHEET_ID}/export?format=xlsx"
OVERVIEW_SHEET   = "Arc Overview"

SHEET_ID = "1M0Aa2p5x7NioaH9-u8FyHq6rH3t5s6Sccs8GoC6pHAM"
CSV_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid=0"

PROPERTIES_URL = "https://raw.githubusercontent.com/one-pace/one-pace-public-subtitles/main/main/title.properties"

def download_excel_file(url, filename, max_retries=3):
    print(f"Downloading latest spreadsheet to {filename}...")
    for attempt in range(1, max_retries + 1):
        try:
            response = requests.get(url, stream=True, timeout=60)
            response.raise_for_status()
            with open(filename, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            print("Download complete!\n")
            return True
        except Exception as e:
            print(f"  [!] Attempt {attempt}/{max_retries} failed: {e}")
            if attempt < max_retries:
                print("  [*] Retrying in 3 seconds...")
                time.sleep(3)
            else:
                print("  [-] All attempts to download the spreadsheet failed.")
                return False


def resolve_prefix(arc_name, arc_map):
    """Map an arc/tab name to its prefix, stripping '(TBR)'/'(WIP)' suffixes."""
    base = re.sub(r"\([^)]*\)", "", str(arc_name)).strip()
    return arc_map.get(clean_string(base))


def parse_episode_label(label):
    """'Egghead 01' -> (1, False); 'Egghead 21 Forward' -> (21, True). None if not an episode."""
    s = str(label).strip()
    if not s or not re.search(r"[A-Za-z]", s):   # require letters -> skips stray total rows
        return None
    low = s.lower()
    is_forward = "forward" in low
    nums = re.findall(r"\d+", re.sub(r"forward", "", low))
    if not nums:
        return None
    return int(nums[-1]), is_forward


# Fixed time-of-day appended to every episode's air date.
RELEASE_TIME_SUFFIX = "T14:15:00.000Z"


def to_iso_release(raw):
    """Convert a spreadsheet 'Release Date' to an ISO-8601 string with the fixed
    time-of-day, or None if the value isn't a clear year-month-day date.

    Accepts a real date/datetime cell, or text 'yyyy.mm.dd' (current format),
    'yyyy-mm-dd', 'yyyy/mm/dd'. Anything else (blank, 'To Be Released', or an
    unexpected layout) returns None, so a format change can only omit a date,
    never produce a wrong one.
    """
    if raw is None:
        return None
    if isinstance(raw, datetime.date):  # covers date and datetime cells
        y, mo, d = raw.year, raw.month, raw.day
    else:
        m = re.fullmatch(r"\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*", str(raw))
        if not m:
            return None
        y, mo, d = (int(x) for x in m.groups())
    try:
        datetime.date(y, mo, d)  # reject impossible dates (e.g. month 13)
    except ValueError:
        return None
    return f"{y:04d}-{mo:02d}-{d:02d}{RELEASE_TIME_SUFFIX}"


def normalize_released(raw):
    """For specials.json: pass a full ISO datetime through as-is, otherwise
    convert a 'yyyy.mm.dd'-style date. None if unrecognized."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}(\.\d+)?Z", s):
        return s
    return to_iso_release(s)


# Longer re-cuts of an episode: "Extended", Muhn's "Alt"/"Fillerver", Skypiea's "Alternate (G-8)".
VARIANT_CUT = re.compile(r"\b(extended|alt|alternate|fillerver)\b", re.I)


def to_runtime_minutes(length):
    """'20:28' -> 20. Minutes may exceed 59 ('60:10'). None if unparseable."""
    m = re.fullmatch(r"\s*(?:(\d{1,3}):)?(\d{1,3}):(\d{2})\s*", str(length or ""))
    if not m:
        return None
    h, mi, s = (int(x or 0) for x in m.groups())
    return ((h * 3600 + mi * 60 + s) + 30) // 60 or None


def build_runtime_index():
    """Map stream file name -> minutes, from the JSONs scr.py writes."""
    out = {}
    for root, _dirs, files in os.walk(os.path.join(BASE_DIR, "stream")):
        for fn in files:
            if not fn.endswith(".json"):
                continue
            try:
                with open(os.path.join(root, fn), "r", encoding="utf-8") as f:
                    streams = json.load(f).get("streams") or []
            except (OSError, ValueError):
                continue
            # Variant cuts run longer; the standard one is what plays by default.
            picks = [s for s in streams if not VARIANT_CUT.search(s.get("releaseName") or "")] or streams
            mins = to_runtime_minutes(picks[0].get("length")) if picks else None
            if mins:
                out[fn[:-5]] = mins
    return out


def runtime_for(vid_id, runtime_index):
    """Meta ids are sometimes 'pp_'-prefixed ('pp_fan_1') while the file is 'fan_1.json'."""
    if not vid_id:
        return None
    return runtime_index.get(vid_id) or runtime_index.get(vid_id[3:] if vid_id.startswith("pp_") else vid_id)


def stamp_season_posters(videos, seasons):
    """Nuvio reads seasonPoster off the first episode in a season that has one,
    so one per season is enough and keeps the file small."""
    posters = {s.get("season"): s.get("poster") for s in (seasons or []) if s.get("poster")}
    done = set()
    for video in videos:
        season = video.get("season")
        if season in posters and season not in done:
            video["seasonPoster"] = posters[season]
            done.add(season)
    return len(done)


def enrich_extra_meta(runtime_index, ratings_map):
    """The other pace meta files are hand-maintained, so enrich them in place."""
    for name in ("pp_muhnpace.json", "pp_onipace.json", "pp_KUMA_SHAVED.json"):
        path = os.path.join(BASE_DIR, "meta", name)
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, ValueError):
            print(f"   [!] {name} missing or unreadable - skipping.")
            continue
        videos = data.get("meta", {}).get("videos") or []
        runtimes = ratings = 0
        for video in videos:
            vid_id = video.get("id", "")
            mins = runtime_for(vid_id, runtime_index)
            if mins:
                video["runtime"] = str(mins)
                runtimes += 1
            if ratings_map.get(vid_id):
                video["rating"] = ratings_map[vid_id]
                ratings += 1
        posters = stamp_season_posters(videos, data.get("meta", {}).get("seasons"))
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        print(f"   - {name:22} {runtimes}/{len(videos)} runtimes, "
              f"{ratings}/{len(videos)} ratings, {posters} season posters")


def collect_release_dates(rows):
    """Return {clean_label: iso_date} for every dated row in an arc tab.

    Captures dates from EVERY row (even ones parse_arc_sheet would skip -- a
    cover-story row with no number, an alternate '(G8)' row that collides with a
    real id) so specials can look up their air date by label later.
    """
    ep_col = header_idx = None
    for i, row in enumerate(rows[:8]):
        for j, c in enumerate(row):
            if c and str(c).strip().lower() == "one pace episode":
                ep_col, header_idx = j, i
                break
        if ep_col is not None:
            break
    if ep_col is None:
        return {}

    release_col = None
    for j, c in enumerate(rows[header_idx]):
        if c and str(c).strip().lower() == "release date":
            release_col = j
            break
    if release_col is None:
        return {}

    out = {}
    for row in rows[header_idx + 1:]:
        if ep_col >= len(row) or row[ep_col] is None or release_col >= len(row):
            continue
        iso = to_iso_release(row[release_col])
        if iso:
            out.setdefault(clean_string(str(row[ep_col])), iso)
    return out


def build_season_map(rows, arc_map):
    """From the Arc Overview rows, build {prefix: season}. Decimal arcs (6.5) fold to int (6)."""
    header_idx = no_col = arc_col = None
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "No." in cells and "Arcs" in cells:
            header_idx, no_col, arc_col = i, cells.index("No."), cells.index("Arcs")
            break
    if header_idx is None:
        return {}

    prefix_to_season = {}
    for row in rows[header_idx + 1:]:
        no_raw  = row[no_col]  if no_col  < len(row) else None
        arc_raw = row[arc_col] if arc_col < len(row) else None
        if no_raw is None or arc_raw is None:
            continue
        if str(arc_raw).strip().lower() == "totals":
            break
        try:
            season = int(float(str(no_raw).strip()))
        except (ValueError, TypeError):
            continue
        prefix = resolve_prefix(arc_raw, arc_map)
        if prefix:
            prefix_to_season[prefix] = season
    return prefix_to_season


def parse_arc_sheet(rows, prefix, season):
    """Build video dicts from one arc tab. A 'Forward' row becomes one 'unreleased' episode."""
    ep_col = header_idx = None
    for i, row in enumerate(rows[:8]):
        for j, c in enumerate(row):
            if c and str(c).strip().lower() == "one pace episode":
                ep_col, header_idx = j, i
                break
        if ep_col is not None:
            break
    if ep_col is None:
        return []

    # Locate the optional "Release Date" column in the same header row.
    release_col = None
    for j, c in enumerate(rows[header_idx]):
        if c and str(c).strip().lower() == "release date":
            release_col = j
            break

    videos, seen = [], set()
    for row in rows[header_idx + 1:]:
        if ep_col >= len(row) or row[ep_col] is None:
            continue
        parsed = parse_episode_label(row[ep_col])
        if not parsed:
            continue
        ep_num, is_forward = parsed
        vid_id = f"{prefix}_{ep_num}"
        if vid_id in seen:
            continue
        seen.add(vid_id)
        video = {
            "season": season,
            "episode": ep_num,
            "id": vid_id,
            "title": "Unreleased" if is_forward else str(row[ep_col]).strip(),
        }
        # Air date from the spreadsheet (skip 'Forward'/unreleased rows).
        if not is_forward and release_col is not None and release_col < len(row):
            iso = to_iso_release(row[release_col])
            if iso:
                video["released"] = iso
                video["firstAired"] = iso
        if is_forward:
            video["_forward"] = True   # temporary flag, removed during the first pass
        videos.append(video)
    # Stremio hides episode 0, so shift a "00"-based arc up by one.    
    if any(v["episode"] == 0 for v in videos):
        for v in videos:
            v["episode"] += 1

    return videos


def build_base_data_from_spreadsheet(arc_map):
    """Replaces the fedew04 JSON: builds meta/videos straight from the official spreadsheet."""
    local_path = os.path.join(BASE_DIR, "one_pace.xlsx")
    # A hint, not a promise: without the file we still download it.
    if "--reuse-spreadsheet" in sys.argv and os.path.exists(local_path):
        print(f"Reusing {local_path} from this run.\n")
    elif not download_excel_file(EPISODE_XLSX_URL, local_path):
        raise RuntimeError("Could not download the One Pace episode spreadsheet.")

    wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
    if OVERVIEW_SHEET not in wb.sheetnames:
        raise RuntimeError(f"'{OVERVIEW_SHEET}' tab not found.")

    overview_rows    = list(wb[OVERVIEW_SHEET].iter_rows(values_only=True))
    prefix_to_season = build_season_map(overview_rows, arc_map)

    all_videos = []
    release_by_label = {}
    for name in wb.sheetnames:
        if name.strip() == OVERVIEW_SHEET:
            continue
        rows = list(wb[name].iter_rows(values_only=True))
        prefix = resolve_prefix(name, arc_map)
        if not prefix:
            print(f"   [!] No ARC_MAP entry for tab '{name}' - skipping.")
            continue

        # Capture air dates from every prefixed tab (even cover-story tabs with
        # no season) so specials can resolve their date by label.
        for lbl, iso in collect_release_dates(rows).items():
            release_by_label.setdefault(lbl, iso)

        season = prefix_to_season.get(prefix)
        if season is None:
            print(f"   [!] No season number for tab '{name}' ({prefix}) - skipping.")
            continue
        arc_videos = parse_arc_sheet(rows, prefix, season)
        print(f"   - {name:34} {prefix:14} S{season:<2}: {len(arc_videos)} eps")
        all_videos.extend(arc_videos)

    wb.close()
    return {
        "meta": {
            "id": "pp_onepacee",
            "type": "series",
            "name": "One Pace",
            "poster":"",
            "logo":"",
            "background":"",
            "description":"",
            "status": "Continuing",
            "videos": all_videos,
        }
    }, release_by_label
    
def clean_string(s):
    """Removes spaces, dashes, apostrophes, and periods for exact matching."""
    return str(s).lower().replace(" ", "").replace("-", "").replace("'", "").replace(".", "")

def get_titles_from_properties(config_arc_map, max_retries=3):
    """Fetches the official titles from the GitHub properties file."""
    print("Fetching official titles from One Pace GitHub...")
    lines = None
    for attempt in range(1, max_retries + 1):
        try:
            req = urllib.request.urlopen(PROPERTIES_URL, timeout=15)
            lines = req.read().decode('utf-8').splitlines()
            break
        except Exception as e:
            print(f"  [!] Attempt {attempt}/{max_retries} failed: {e}")
            if attempt < max_retries:
                print("  [*] Retrying in 3 seconds...")
                time.sleep(3)

    if lines is None:
        raise RuntimeError("Could not fetch title.properties.")

    id_to_title = {}
    key_to_title = {}

    for line in lines:
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, title = line.split("=", 1)
        
        # Clean the title: Strip whitespace and replace all double quotes with single quotes
        title = title.strip().replace('\\"', "'").replace('"', "'")
        
        if not title: 
            continue

        # e.g., "specials_04": "One Piece Fan Letter"
        original_key = key.split(".")[0]
        key_to_title[original_key] = title 

        raw_prefix, raw_ep = original_key.split("_")
        ep_num_int = int(raw_ep)

        clean_prefix = raw_prefix.lower().replace(" ", "")
        mapped_prefix = config_arc_map.get(clean_prefix, raw_prefix.upper())
        video_id = f"{mapped_prefix}_{ep_num_int}"

        id_to_title[video_id] = title

    return id_to_title, key_to_title

def main():
    print("1. Building base episode list from the official One Pace spreadsheet...")
    try:
        data, release_by_label = build_base_data_from_spreadsheet(ARC_PREFIXES)
    except Exception as e:
        print(f"Failed to build base data: {e}")
        return

    print("2. Fetching descriptions from Google Sheets...")
    try:
        req = urllib.request.urlopen(CSV_URL, timeout=10)
        csv_content = req.read().decode('utf-8')
    except Exception as e:
        print(f"Failed to download spreadsheet: {e}")
        return

    # Build description dictionaries
    descriptions_map = {}
    title_to_desc = {}   # Fallback map based on exact English Title
    title_en_map = {}    # video_id -> official English title from the descriptions sheet

    reader = csv.DictReader(StringIO(csv_content))
    for row in reader:
        arc_title = row.get("arc_title", "").strip()
        arc_part = row.get("arc_part", "").strip()
        title_en = row.get("title_en", "").strip()
        desc = row.get("description_en", "").strip()

        # Resolve this row's video_id once (used for both the title and description maps)
        row_video_id = None
        if arc_title and arc_part:
            clean_arc = clean_string(arc_title)
            if clean_arc in ARC_PREFIXES:
                matched_prefix = ARC_PREFIXES[clean_arc]
                try:
                    arc_part_str = str(int(arc_part))
                except ValueError:
                    arc_part_str = arc_part
                row_video_id = f"{matched_prefix}_{arc_part_str}"

        # Title map: capture title_en even when this row has no description
        if row_video_id and title_en:
            title_en_map[row_video_id] = title_en

        if not desc:
            continue

        # 1. Save by EXACT title (cleaning quotes to match properties file)
        if title_en:
            clean_title_en = title_en.replace('\\"', "'").replace('"', "'")
            title_to_desc[clean_title_en] = desc

        # 2. Save by ID mapping
        if row_video_id:
            descriptions_map[row_video_id] = desc

    print("3. Getting official titles from properties...")
    try:
        titles_map, key_to_title = get_titles_from_properties(ARC_PREFIXES)
    except Exception as e:
        # Without these, every episode falls back to a spreadsheet label.
        print(f"Failed to fetch properties: {e}")
        sys.exit(1)

    print("4. Loading specials.json configuration...")
    custom_titles = {}
    custom_overviews = {}
    try:
        SPECIALS_PATH = os.path.join(BASE_DIR, 'meta', 'specials.json')
        with open(SPECIALS_PATH, 'r', encoding='utf-8') as f:
            specials_config = json.load(f)

        # Pull out optional custom-name overrides BEFORE iterating specials,
        # so the loop below only sees real special entries (each with an "id").
        # A custom_names value is either a plain title string, or an object
        # {"title": ..., "overview": ...} to also set a description.
        for cid, cval in specials_config.pop("custom_names", {}).items():
            if isinstance(cval, dict):
                if cval.get("title"):
                    custom_titles[cid] = cval["title"]
                overview = cval.get("overview") or cval.get("description")
                if overview:
                    custom_overviews[cid] = overview
            else:
                custom_titles[cid] = cval

        specials_by_id = {}
        for spec_key, spec_val in specials_config.items():
            vid_id = spec_val["id"]
            spec_val["original_key"] = spec_key  # Keep track of the key (e.g. specials_04)
            specials_by_id[vid_id] = spec_val

    except FileNotFoundError:
        print("specials.json not found. Proceeding without special reordering.")
        specials_by_id = {}
    
    print("4.5. Loading thumbnails.json configuration...")
    try:
        THUMBNAILS_PATH = os.path.join(BASE_DIR, 'meta', 'thumbnails.json')
        with open(THUMBNAILS_PATH, 'r', encoding='utf-8') as f:
            thumbnails_map = json.load(f)
    except FileNotFoundError:
        print("thumbnails.json not found. Proceeding without custom thumbnails.")
        thumbnails_map = {}
        
    print("4.6. Loading ratings.json configuration...")
    try:
        RATINGS_PATH = os.path.join(BASE_DIR, 'meta', 'ratings.json')
        with open(RATINGS_PATH, 'r', encoding='utf-8') as f:
            ratings_map = json.load(f)
    except FileNotFoundError:
        print("ratings.json not found. Proceeding without ratings.")
        ratings_map = {}

    print("5. Applying transformations, titles, and injecting specials...")
    meta = data.get("meta", {})
    
    # Static Data Injection
    meta.update({
        "id": "pp_onepacee",
        "poster": "https://cdn.jsdelivr.net/gh/6ip/onepace-assets-prm@main/public/poster.jpg",
        "background": "https://cdn.jsdelivr.net/gh/6ip/onepace-assets-prm@main/public/background_pace.jpg",
        "logo": "https://cdn.jsdelivr.net/gh/6ip/onepace-assets-prm@main/public/logo.png",
        "description": "Experience One Piece without the filler. This manga-accurate cut removes padded scenes, saving you hundreds of hours while staying true to Oda's original vision.",
        "releaseInfo": "1999-",
        "year": "1999-",
        "imdbRating": "9.0",
        "country": "Japan",
        "language": "Japanese, English",
        "ageRating": "TV-14",
        "awards": "7 wins & 21 nominations total",
        "released": "1999-10-20T00:00:00.000Z",
        "genres": ["Animation", "Action", "Adventure"],
        "writer": ["Eiichirô Oda"],
        "cast": [
            "Mayumi Tanaka", "Kazuya Nakai", "Akemi Okamura", "Kappei Yamaguchi",
            "Hiroaki Hirata", "Ikue Otani", "Yuriko Yamaguchi", "Kazuki Yao",
            "Cho", "Katsuhisa Hōki"
        ],
        "trailers": [{"source": "1KMcoJBMWE4", "type": "Trailer", "name": "One Piece Trailer"}],
        "trailerStreams": [{"title": "One Piece Trailer", "ytId": "1KMcoJBMWE4"}],
        "app_extras": {
            "cast": [
                {"name": "Mayumi Tanaka", "character": "Monkey D. Luffy", "photo": "https://media.themoviedb.org/t/p/w132_and_h132_face/by4t1tYtEXsfbFj9TvOjozBmQla.jpg"},
                {"name": "Kazuya Nakai", "character": "Roronoa Zoro", "photo": "https://media.themoviedb.org/t/p/w132_and_h132_face/cOHSa0FBrG3u9P6g8A56sInkvod.jpg"},
                {"name": "Akemi Okamura", "character": "Nami", "photo": "https://media.themoviedb.org/t/p/w132_and_h132_face/qEwVfsrA23SCzRZMSlghw2sHctQ.jpg"},
                {"name": "Kappei Yamaguchi", "character": "Usopp", "photo": "https://media.themoviedb.org/t/p/w132_and_h132_face/mJyxKRZxLv9D7LH5KcNSkjSKYOB.jpg"},
                {"name": "Hiroaki Hirata", "character": "Sanji", "photo": "https://media.themoviedb.org/t/p/w132_and_h132_face/mMWEdlssJm3TVuXW4Wb7pQeX20Z.jpg"},
                {"name": "Ikue Otani", "character": "Tony Tony Chopper", "photo": "https://media.themoviedb.org/t/p/w132_and_h132_face/my8LBLQ4MsK4hRz1PAATIqtieaI.jpg"},
                {"name": "Yuriko Yamaguchi", "character": "Nico Robin", "photo": "https://media.themoviedb.org/t/p/w132_and_h132_face/14hV8oOMkZddPpvdpVSODRwVPgv.jpg"},
                {"name": "Kazuki Yao", "character": "Franky", "photo": "https://media.themoviedb.org/t/p/w132_and_h132_face/kjETB1FaDaGozGjdq4Hu1UjrQDf.jpg"},
                {"name": "Cho", "character": "Brook", "photo": "https://media.themoviedb.org/t/p/w132_and_h132_face/70OZtKOHEGvbuMQAbQ01cGxrUDO.jpg"},
                {"name": "Katsuhisa Hōki", "character": "Jinbe", "photo": "https://media.themoviedb.org/t/p/w132_and_h132_face/stuUwLrSOmQhOmNpO4kCrlu3ALF.jpg"}
            ],
            "writers": [
                {"name": "Eiichirô Oda", "character": "Original Manga", "photo": "https://media.themoviedb.org/t/p/w132_and_h132_face/4YkEwEiz0E7EKYZzW8Fy4EGTITO.jpg"}
            ]
        }
    })
    
    meta["links"] = [
        {"name": "9.0", "category": "imdb", "url": "https://imdb.com/title/tt0388629"},
        {"name": "Animation", "category": "Genres", "url": "stremio:///discover/https%3A%2F%2Fv3-cinemeta.strem.io%2Fmanifest.json/series/top?genre=Animation"},
        {"name": "Action", "category": "Genres", "url": "stremio:///discover/https%3A%2F%2Fv3-cinemeta.strem.io%2Fmanifest.json/series/top?genre=Action"},
        {"name": "Adventure", "category": "Genres", "url": "stremio:///discover/https%3A%2F%2Fv3-cinemeta.strem.io%2Fmanifest.json/series/top?genre=Adventure"},
        {"name": "Eiichirô Oda", "category": "Writers", "url": "stremio:///search?search=Eiichir%C3%B4%20Oda"},
        {"name": "Mayumi Tanaka", "category": "Cast", "url": "stremio:///search?search=Mayumi%20Tanaka"},
        {"name": "Kazuya Nakai", "category": "Cast", "url": "stremio:///search?search=Kazuya%20Nakai"},
        {"name": "Akemi Okamura", "category": "Cast", "url": "stremio:///search?search=Akemi%20Okamura"},
        {"name": "Kappei Yamaguchi", "category": "Cast", "url": "stremio:///search?search=Kappei%20Yamaguchi"},
        {"name": "Hiroaki Hirata", "category": "Cast", "url": "stremio:///search?search=Hiroaki%20Hirata"},
        {"name": "Ikue Otani", "category": "Cast", "url": "stremio:///search?search=Ikue%20Otani"},
        {"name": "Yuriko Yamaguchi", "category": "Cast", "url": "stremio:///search?search=Yuriko%20Yamaguchi"},
        {"name": "Kazuki Yao", "category": "Cast", "url": "stremio:///search?search=Kazuki%20Yao"},
        {"name": "Cho", "category": "Cast", "url": "stremio:///search?search=Cho"},
        {"name": "Katsuhisa Hōki", "category": "Cast", "url": "stremio:///search?search=Katsuhisa%20H%C5%8Dki"}
    ]

    meta["seasons"] = [
        {
            "season": i,
            "poster": f"https://cdn.jsdelivr.net/gh/6ip/onepace-assets-prm@main/public/poster-s/poster-s{str(i).zfill(2)}.jpg"
        } for i in range(0, TOTAL_SEASONS + 1)
    ]

    desc_count = 0
    normal_videos = []
    special_videos_extracted = []

    # FIRST PASS: Clean IDs, override titles, set descriptions, filter configured specials
    for video in meta.get("videos", []):
        vid_id = video.get("id", "")
        if not str(vid_id).startswith("pp_"):
            vid_id = f"{vid_id}"
            video["id"] = vid_id
            
        # Title priority:
        #   forward (unreleased) episodes  -> always "Unreleased"
        #   released episodes -> title.properties > descriptions-sheet title_en > custom_names > spreadsheet label
        if video.pop("_forward", False):
            video["title"] = "Unreleased"
        elif vid_id in titles_map:
            video["title"] = titles_map[vid_id]
        elif vid_id in title_en_map:
            video["title"] = title_en_map[vid_id]
        elif vid_id in custom_titles:
            video["title"] = custom_titles[vid_id]
        # else: keep the builder's fallback label (e.g. "Long Ring Long Land 00")

        season_num = video.get("season")
        
        # --- Thumbnail priority (normal videos) ---
        if vid_id in thumbnails_map:
            video["thumbnail"] = thumbnails_map[vid_id]
        elif season_num and 1 <= season_num <= TOTAL_SEASONS:
            s_padded = str(season_num).zfill(2)
            video["thumbnail"] = f"https://images.weserv.nl/?url=cdn.jsdelivr.net/gh/6ip/onepace-assets-prm@main/public/poster-s/poster-s{s_padded}.jpg&w=1280&h=720&fit=cover&a=center"
        else:
            video["thumbnail"] = "https://image.tmdb.org/t/p/w500/iN5LKyvyWUWwqbjaQfKFXoo8mch.jpg"

        # Inject Descriptions (Try ID map first, then Title map)
        clean_vid_title = video.get("title", "").replace('\\"', "'").replace('"', "'")
        
        if vid_id in descriptions_map:
            video["description"] = descriptions_map[vid_id]
            video["overview"] = descriptions_map[vid_id]
            desc_count += 1
        elif clean_vid_title in title_to_desc:
            video["description"] = title_to_desc[clean_vid_title]
            video["overview"] = title_to_desc[clean_vid_title]
            desc_count += 1
        elif vid_id in custom_overviews:
            video["description"] = custom_overviews[vid_id]
            video["overview"] = custom_overviews[vid_id]
            desc_count += 1

        if vid_id in specials_by_id:
            special_videos_extracted.append(video)
        else:
            normal_videos.append(video)

    # CONFIGURED SPECIALS: Create any missing specials defined in specials.json
    found_spec_ids = {v["id"] for v in special_videos_extracted}
    processed_original_keys = set()
    
    for spec_id, spec_val in specials_by_id.items():
        orig_key = spec_val["original_key"]
        processed_original_keys.add(orig_key)
        
        if spec_id not in found_spec_ids:
            spec_title = spec_val.get("custom_title", key_to_title.get(orig_key, "Special Episode"))
            # Try ID, then Title fallback, then custom_overview from specials.json
            desc = descriptions_map.get(spec_id) or title_to_desc.get(spec_title) or spec_val.get("custom_overview", "")
            if desc:
                desc_count += 1
                
            new_special = {
                "id": spec_id,
                "title": spec_title,
                "description": desc,
                "overview": desc,
                "thumbnail": "https://image.tmdb.org/t/p/w500/iN5LKyvyWUWwqbjaQfKFXoo8mch.jpg"
            }
            special_videos_extracted.append(new_special)

    # UNCONFIGURED SPECIALS AUTO-CATCHER
    # Finds any special in title.properties not mentioned in specials.json and auto-adds to Season 0
    all_property_keys = key_to_title.keys()
    for prop_key in all_property_keys:
        if prop_key.startswith("specials_") and prop_key not in processed_original_keys:
            raw_prefix, raw_ep = prop_key.split("_")
            fallback_id = f"{raw_prefix.upper()}_{int(raw_ep)}"
            spec_title = key_to_title[prop_key]
            
            # Fetch description strictly by exact title matching
            desc = title_to_desc.get(spec_title, "")
            if desc:
                desc_count += 1

            unconfigured_special = {
                "id": fallback_id,
                "title": f"[Special] {spec_title}",
                "description": desc,
                "overview": desc,
                "season": 0, # Send unconfigured straight to Season 0
                "episode": int(raw_ep),
                "thumbnail": thumbnails_map.get(fallback_id, "https://image.tmdb.org/t/p/w500/iN5LKyvyWUWwqbjaQfKFXoo8mch.jpg"),
            }
            normal_videos.append(unconfigured_special)

    # SECOND PASS: Re-inject the configured specials at the end of their assigned seasons
    for spec_vid in special_videos_extracted:
        vid_id = spec_vid["id"]
        target_season = specials_by_id[vid_id]["season"]
        title_prefix = specials_by_id[vid_id]["title_start"]
        custom_thumb = specials_by_id[vid_id].get("thum")

        original_title = spec_vid.get("title", "")
        if not original_title.startswith(title_prefix.strip()):
            spec_vid["title"] = f"{title_prefix}{original_title}"

        spec_vid["season"] = target_season

        # Air date, in priority:
        #   1) explicit "released" in specials.json (date or full ISO)
        #   2) "release_label" -> the exact spreadsheet row's date
        #   3) auto: the special's own title matches a spreadsheet row label
        #      (covers cover-story specials whose title == the sheet label)
        rel = normalize_released(specials_by_id[vid_id].get("released"))
        if not rel:
            label_hint = specials_by_id[vid_id].get("release_label")
            if label_hint:
                rel = release_by_label.get(clean_string(label_hint))
        if not rel and original_title:
            rel = release_by_label.get(clean_string(original_title))
        if rel:
            spec_vid["released"] = rel
            spec_vid["firstAired"] = rel

        # --- 4-tier thumbnail priority (specials) ---
        if vid_id in thumbnails_map:
            # Priority 1: Custom thumbnail from thumbnails.json
            spec_vid["thumbnail"] = thumbnails_map[vid_id]
        elif custom_thumb:
            # Priority 2: Custom thumbnail from specials.json
            spec_vid["thumbnail"] = custom_thumb
        elif 1 <= target_season <= TOTAL_SEASONS:
            # Priority 3: Season-specific poster
            s_padded = str(target_season).zfill(2)
            spec_vid["thumbnail"] = f"https://images.weserv.nl/?url=cdn.jsdelivr.net/gh/6ip/onepace-assets-prm@main/public/poster-s/poster-s{s_padded}.jpg&w=1280&h=720&fit=cover&a=center"
        else:
            # Priority 4: Fallback TMDB background
            spec_vid["thumbnail"] = "https://image.tmdb.org/t/p/w500/iN5LKyvyWUWwqbjaQfKFXoo8mch.jpg"

        # Apply custom_overview if no description was set from the spreadsheet
        custom_overview = specials_by_id[vid_id].get("custom_overview")
        if custom_overview and not spec_vid.get("description"):
            spec_vid["description"] = custom_overview
            spec_vid["overview"] = custom_overview

        # Find the max episode number currently existing in the target season
        episodes_in_target = [v.get("episode", 0) for v in normal_videos if v.get("season") == target_season]
        highest_ep = max(episodes_in_target) if episodes_in_target else 0

        spec_vid["episode"] = highest_ep + 1
        normal_videos.append(spec_vid)

    # Final Sort to ensure the UI looks clean
    normal_videos.sort(key=lambda x: (x.get("season", 0), x.get("episode", 0)))
    meta["videos"] = normal_videos

    print("5.5. Adding runtimes from stream/ and ratings from ratings.json...")
    runtime_index = build_runtime_index()
    runtime_count = 0
    rating_count = 0
    for video in normal_videos:
        vid_id = video.get("id", "")
        mins = runtime_for(vid_id, runtime_index)
        if mins:
            video["runtime"] = str(mins)
            runtime_count += 1
        # ratings.json is keyed by the exact meta id, so no prefix juggling here.
        if ratings_map.get(vid_id):
            video["rating"] = ratings_map[vid_id]
            rating_count += 1

    poster_count = stamp_season_posters(normal_videos, meta.get("seasons"))

    print("6. Saving fully assembled JSON...")
    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

    print("7. Adding runtimes and ratings to the other pace meta files...")
    enrich_extra_meta(runtime_index, ratings_map)

    print(f"\nDone! Saved to {OUTPUT_JSON}.")
    print(f"Descriptions matched and injected: {desc_count}")
    print(f"Runtimes added from stream/: {runtime_count}")
    print(f"Ratings added from ratings.json: {rating_count}")
    print(f"Season posters stamped: {poster_count}")
    print(f"Specials safely reordered/injected: {len(special_videos_extracted)}")

if __name__ == "__main__":
    main()